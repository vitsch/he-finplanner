"""
Recalibrated POMDP — Russell Group Panel  (c_sum_18)
Real data integration: HESA enrolments (data_1), HESA Finance Statistics (data_2),
Home Office visa statistics (data_3), ONS National Population Projections (data_3).

Data status:
  Step 1  HESA Table 1 (data_1/table-1.zip)              → T_pooled   [REAL]
  Step 2  HESA enrolment % changes (data_1/table-1.zip)   → Omega      [REAL — HESA-based (c_sum_18)]
  Step 3  HESA Finance Statistics (data_2/table-1.csv)    → R_net      [REAL — clean window 2004/05-2017/18]
  Step 4  ONS 2024-based NPP (data_3/ons_npp_...)         → b0 base    [REAL — Ages 15-29, 2024-2034 (+0.31%/yr)]
  Step 5  Home Office visa stats (data_3/...)             → b0 visa    [REAL]
  Step 6  Monte Carlo comparison (6 strategies)                        [REAL parameters]
"""

import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from hmmlearn import hmm
import warnings, os, zipfile, csv, io
warnings.filterwarnings('ignore')

WD = os.path.dirname(os.path.abspath(__file__))
os.chdir(WD)

DATA_FILES = {
    "hesa_enrolments":   "data_1/table-1.zip",
    "hesa_finance_csv":  "data_2/table-1.csv",
    "ho_visas":          "data_3/home_office_edu_visas_sep2025.xlsx",
    "ons_npp":           "data_3/ons_npp_2024based_england_summary.xlsx",
    "ucas_eoc":          "ucas_eoc_provider_level.csv",       # unavailable — synthetic fallback
}

# ── Institutions (c_sum_5 §1.2) — UKPRNs added for real data loading
INSTITUTIONS = {
    "Warwick": {"ucas": "W20", "ukprn": "10007163", "region": "Midlands"},
    "Exeter":  {"ucas": "E14", "ukprn": "10007792", "region": "South West"},
    "Leeds":   {"ucas": "L23", "ukprn": "10007795", "region": "North"},
    "Bristol": {"ukprn": "10007786", "ucas": "B78",  "region": "South West"},
}
INST_NAMES  = list(INSTITUTIONS.keys())
UKPRN_MAP   = {v["ukprn"]: k for k, v in INSTITUTIONS.items()}
N_INST      = len(INST_NAMES)

# Extended panel (c_sum_21): Durham / Nottingham / Sheffield added for T_pooled robustness
INSTITUTIONS_EXT = dict(INSTITUTIONS)
INSTITUTIONS_EXT.update({
    "Durham":     {"ucas": "D86", "ukprn": "10007143", "region": "North East"},
    "Nottingham": {"ucas": "N84", "ukprn": "10007154", "region": "East Midlands"},
    "Sheffield":  {"ucas": "S18", "ukprn": "10007157", "region": "Yorkshire"},
})
UKPRN_MAP_EXT  = {v["ukprn"]: k for k, v in INSTITUTIONS_EXT.items()}
INST_NAMES_EXT = list(INSTITUTIONS_EXT.keys())

# ── Year ranges
# Real HESA enrolment data: 2014/15–2024/25 → 11 data points → 10 % change transitions
ACAD_YEARS  = [f"{y}/{str(y+1)[2:]}" for y in range(2014, 2025)]  # ['2014/15', …, '2024/25']
YEARS       = list(range(2014, 2025))   # integer axis for plots (start-year of each acad year)
PCT_YEARS   = ACAD_YEARS[1:]            # ['2015/16', …, '2024/25'] (10 % change observations)
CYCLES      = list(range(2013, 2024))   # UCAS EOC cycles 2013–2023

# ── R_net calibration constants (c_sum_12 clean window 2004/05–2017/18)
# Surplus % observations by regime, verified against HESA Finance Plus (data_2)
# and HESA Finance Statistics CSV (data_2/table-1.csv); USS distortion years excluded
CLEAN_SURPLUS_PCT = {
    0: [0.3, 1.4, -2.0, -2.0, 1.0, 1.4, -1.2, -0.8, 4.6, 1.1, 1.7, 1.5, 1.7, -0.6, 1.5],   # Low
    1: [2.7, 3.8, 2.4, 3.6, 7.0, 8.9, 1.8, 3.7, 5.9, 0.9, 2.5, 3.9, 6.6, 1.5,             # Stable
        1.9, 2.3, 2.7, 3.7, 6.4, 10.5],
    2: [4.8, 6.8, 9.5, 9.7, 4.0, 7.0, 10.0, 9.7, 4.2, 9.3, 7.2],                           # Growth
}

# ── POMDP constants
N_STATES  = 3;  LOW, STABLE, GROWTH = 0, 1, 2
N_ACTIONS = 5;  CONSERVE, MAINTAIN, RECRUIT, INVEST, RATIONALISE = 0, 1, 2, 3, 4
N_OBS     = 4;  POOR, WEAK, MODERATE, STRONG = 0, 1, 2, 3
ACTION_NAMES = ["Conserve", "Maintain", "Recruit", "Invest", "Rationalise"]
STATE_NAMES  = ["Low", "Stable", "Growth"]

GAMMA     = 0.95
T_HORIZON = 12
DELTA     = 0.10
LAMBDA_V  = 1.5
MU_OV     = 1.2
N_TRIALS  = 500
N_SAT     = 30
PER_STEP_CRISIS = -8
U_MIN = -25.0


# ============================================================
# STEP 1 — HESA TABLE 1 (data_1/table-1.zip) → POOLED HMM → T
# ============================================================

def load_hesa_enrolments():
    """Load real HESA enrolment data from data_1/table-1.zip.
    Returns dict {inst: np.array of annual FTE totals, 2014/15–2024/25 (11 points)}.
    Filter: UKPRN in panel, Entrant=All, Level=All, Mode=All, Category=Total/Total.
    """
    path = DATA_FILES["hesa_enrolments"]
    if os.path.exists(path):
        data = {inst: {} for inst in INST_NAMES}
        with zipfile.ZipFile(path) as z:
            for fname in sorted(z.namelist()):
                if not fname.endswith('.csv'):
                    continue
                # filename: table-1-(2014-15).csv → acad year '2014/15'
                yr_raw = fname.replace('table-1-(', '').replace(').csv', '')
                acad_yr = yr_raw.replace('-', '/')      # '2014/15'
                with z.open(fname) as f:
                    lines = f.read().decode('utf-8-sig', errors='replace').split('\n')
                # header is at row 13 (0-indexed)
                hdr_idx = next(i for i, l in enumerate(lines) if l.startswith('UKPRN'))
                reader = csv.reader(io.StringIO('\n'.join(lines[hdr_idx:])))
                next(reader)  # skip header
                for row in reader:
                    if len(row) < 11:
                        continue
                    ukprn = row[0].strip()
                    if ukprn not in UKPRN_MAP:
                        continue
                    entrant  = row[4].strip()
                    level    = row[5].strip()
                    mode     = row[6].strip()
                    cat_type = row[8].strip() if len(row) > 8 else ''
                    cat_val  = row[9].strip() if len(row) > 9 else ''
                    value    = row[10].strip() if len(row) > 10 else ''
                    if entrant == 'All' and level == 'All' and mode == 'All' \
                            and cat_type == 'Total' and cat_val == 'Total':
                        try:
                            data[UKPRN_MAP[ukprn]][acad_yr] = int(value.replace(',', ''))
                        except ValueError:
                            pass
        # Build ordered arrays aligned to ACAD_YEARS
        result = {}
        for inst in INST_NAMES:
            arr = np.array([data[inst].get(yr, np.nan) for yr in ACAD_YEARS], dtype=float)
            result[inst] = arr
        print("  [REAL DATA] Loaded HESA enrolments from", path)
        print(f"  Academic years: {ACAD_YEARS[0]} – {ACAD_YEARS[-1]}  ({len(ACAD_YEARS)} points)")
        for inst in INST_NAMES:
            vals = result[inst]
            print(f"    {inst:8s}: {int(vals[0]):,} → {int(vals[-1]):,}  "
                  f"({'all present' if not np.any(np.isnan(vals)) else 'some missing'})")
        return result

    print("  [SYNTHETIC] HESA enrolment ZIP not found — using documented synthetic data")
    rng = np.random.default_rng(seed=0)
    def make_series(base, growth_rates, noise_sd=0.005):
        vals = [base]
        for g in growth_rates:
            vals.append(vals[-1] * (1 + g + rng.normal(0, noise_sd)))
        return np.array(vals)
    return {
        "Warwick": make_series(23685, [.041,.015,.026,.015,.029,.048,.025,.009,.012,-.053]),
        "Exeter":  make_series(20555, [.054,.069,.038,.040,.077,.123,.074,.011,.002,.007]),
        "Leeds":   make_series(31030, [.024,.047,.034,.053,.002,.014,.009,.006,.018,-.092]),
        "Bristol": make_series(21555, [.016,.077,.053,.044,.055,.088,.057,.019,.001,.009]),
    }


def load_hesa_enrolments_extended():
    """Load HESA enrolments for the extended 7-institution panel (c_sum_21).
    Adds Durham (10007143), Nottingham (10007154), Sheffield (10007157)
    to the core 4-institution panel for T_pooled robustness validation.
    Uses the same data_1/table-1.zip; no new download needed.
    """
    path = DATA_FILES["hesa_enrolments"]
    data = {inst: {} for inst in INST_NAMES_EXT}
    if os.path.exists(path):
        with zipfile.ZipFile(path) as z:
            for fname in sorted(z.namelist()):
                if not fname.endswith('.csv'):
                    continue
                yr_raw  = fname.replace('table-1-(', '').replace(').csv', '')
                acad_yr = yr_raw.replace('-', '/')
                with z.open(fname) as f:
                    lines = f.read().decode('utf-8-sig', errors='replace').split('\n')
                hdr_idx = next(i for i, l in enumerate(lines) if l.startswith('UKPRN'))
                reader = csv.reader(io.StringIO('\n'.join(lines[hdr_idx:])))
                next(reader)
                for row in reader:
                    if len(row) < 11:
                        continue
                    ukprn = row[0].strip()
                    if ukprn not in UKPRN_MAP_EXT:
                        continue
                    entrant = row[4].strip(); level = row[5].strip()
                    mode    = row[6].strip()
                    cat_type = row[8].strip() if len(row) > 8 else ''
                    cat_val  = row[9].strip() if len(row) > 9 else ''
                    value    = row[10].strip() if len(row) > 10 else ''
                    if entrant == 'All' and level == 'All' and mode == 'All' \
                            and cat_type == 'Total' and cat_val == 'Total':
                        try:
                            data[UKPRN_MAP_EXT[ukprn]][acad_yr] = int(value.replace(',', ''))
                        except ValueError:
                            pass
        result = {}
        for inst in INST_NAMES_EXT:
            arr = np.array([data[inst].get(yr, np.nan) for yr in ACAD_YEARS], dtype=float)
            result[inst] = arr
        new_insts = [i for i in INST_NAMES_EXT if i not in INST_NAMES]
        print(f"  [REAL DATA] Extended panel loaded ({len(INST_NAMES_EXT)} institutions):")
        for inst in new_insts:
            vals = result[inst]
            ok = 'all present' if not np.any(np.isnan(vals)) else 'some missing'
            print(f"    {inst:10s}: {int(vals[0]):,} → {int(vals[-1]):,}  ({ok})")
        return result
    # Fallback: extend existing enrolments with synthetic data for new institutions
    print("  [SYNTHETIC] Extended panel: ZIP not found, using synthetic for new institutions")
    base = load_hesa_enrolments()
    rng = np.random.default_rng(seed=7)
    base["Durham"]     = rng.normal(18000, 500, len(ACAD_YEARS)).cumsum()
    base["Nottingham"] = rng.normal(33000, 700, len(ACAD_YEARS)).cumsum()
    base["Sheffield"]  = rng.normal(28000, 600, len(ACAD_YEARS)).cumsum()
    return base


def compute_pct_change(enrolments):
    return {inst: np.diff(arr) / arr[:-1] * 100 for inst, arr in enrolments.items()}


def fit_pooled_hmm(pct_changes):
    series_list = [pct_changes[inst] for inst in INST_NAMES]
    X = np.concatenate(series_list).reshape(-1, 1)
    lengths = [len(s) for s in series_list]
    model = hmm.GaussianHMM(n_components=3, covariance_type="full",
                             n_iter=500, random_state=42, tol=1e-4)
    model.fit(X, lengths)
    order = np.argsort(model.means_.flatten())
    T_raw      = model.transmat_[np.ix_(order, order)]
    means_     = model.means_.flatten()[order]
    covars_    = model.covars_.flatten()[order]
    startprob_ = model.startprob_[order]
    state_seqs = {}
    pos = 0
    for inst, length in zip(INST_NAMES, lengths):
        seg = X[pos:pos+length]
        _, raw_states = model.decode(seg)
        state_seqs[inst] = np.array([order.tolist().index(s) for s in raw_states])
        pos += length
    return T_raw, means_, covars_, startprob_, state_seqs, model


def compute_empirical_thresholds(pct_changes):
    all_deltas = np.concatenate(list(pct_changes.values()))
    return np.percentile(all_deltas, 33), np.percentile(all_deltas, 67)


def assign_states_threshold(pct_changes, low_thresh, high_thresh, inst_names=None):
    """Assign regime labels by 33rd/67th percentile thresholds.
    GaussianHMM on a 10-transition panel is prone to degenerate local minima;
    the threshold approach is transparent, reproducible, and matches c_sum_10.
    inst_names defaults to INST_NAMES; pass INST_NAMES_EXT for the extended panel.
    """
    if inst_names is None:
        inst_names = INST_NAMES
    state_seqs = {}
    for inst in inst_names:
        states = []
        for delta in pct_changes[inst]:
            if delta < low_thresh:
                states.append(LOW)
            elif delta >= high_thresh:
                states.append(GROWTH)
            else:
                states.append(STABLE)
        state_seqs[inst] = np.array(states)
    return state_seqs


def compute_T_empirical(state_seqs, n_states=N_STATES, laplace=1.0, inst_names=None):
    """Empirical transition matrix with Laplace smoothing (add-1).
    Counts all consecutive (s_t, s_{t+1}) pairs across all institutions.
    inst_names defaults to INST_NAMES; pass INST_NAMES_EXT for the extended panel.
    """
    if inst_names is None:
        inst_names = INST_NAMES
    counts = np.ones((n_states, n_states)) * laplace
    for inst in inst_names:
        seq = state_seqs[inst]
        for i in range(len(seq) - 1):
            counts[seq[i], seq[i + 1]] += 1
    T = counts / counts.sum(axis=1, keepdims=True)
    return T


# ============================================================
# STEP 2 — UCAS END-OF-CYCLE: APPLICATION SIGNALS → OMEGA
# (synthetic — provider-level data requires UCAS account)
# ============================================================

def load_ucas_eoc():
    path = DATA_FILES["ucas_eoc"]
    if os.path.exists(path):
        import csv
        data = {inst: {} for inst in INST_NAMES}
        with open(path) as f:
            reader = csv.DictReader(f)
            for row in reader:
                for inst in INST_NAMES:
                    if row.get("provider_code") == INSTITUTIONS[inst]["ucas"]:
                        yr = int(row["cycle_year"])
                        data[inst][yr] = (int(row["applications_jan"]), int(row["placed"]))
        print("  [REAL DATA] Loaded UCAS End-of-Cycle from", path)
        return data

    print("  [SYNTHETIC] UCAS EOC — provider account required; using documented synthetic data")
    rng = np.random.default_rng(seed=1)
    base_apps   = {"Warwick": 38000, "Exeter": 42000, "Leeds": 55000, "Bristol": 44000}
    base_placed = {"Warwick":  5200, "Exeter":  6200, "Leeds": 10500, "Bristol":  7000}
    data = {inst: {} for inst in INST_NAMES}
    for inst in INST_NAMES:
        apps_t, plac_t = base_apps[inst], base_placed[inst]
        for yr in CYCLES:
            g_apps = rng.normal(0.015, 0.02)
            g_plac = rng.normal(0.018, 0.02)
            if yr == 2021: g_apps += 0.06; g_plac += 0.04
            if yr == 2023 and inst in ["Warwick", "Leeds"]: g_plac -= 0.10
            apps_t = int(apps_t * (1 + g_apps))
            plac_t = int(plac_t * (1 + g_plac))
            data[inst][yr] = (apps_t, plac_t)
    return data


def build_omega_hesa(pct_changes, state_seqs, low_thresh, high_thresh):
    """
    Build emission matrix Ω from real HESA enrolment % changes (c_sum_18).

    Observation thresholds (c_sum_16 §4.3):
      Poor:     Δ% < 0                   (actual enrolment decline)
      Weak:     0 ≤ Δ% < low_thresh      (slow positive growth, still Low regime)
      Moderate: low_thresh ≤ Δ% < high_thresh  (coincident with Stable regime)
      Strong:   Δ% ≥ high_thresh              (coincident with Growth regime)

    The Moderate/Strong thresholds are coincident with the state boundaries,
    so Stable and Growth rows are near-deterministic. The Low row carries
    real signal: Poor = actual decline (2/13 Low obs), Weak = slow growth.
    Laplace-1 smoothing prevents zero probabilities on off-diagonal cells.
    """
    raw = np.zeros((N_STATES, N_OBS), dtype=int)
    for inst in INST_NAMES:
        for delta, s in zip(pct_changes[inst], state_seqs[inst]):
            if delta < 0.0:
                o = POOR
            elif delta < low_thresh:
                o = WEAK
            elif delta < high_thresh:
                o = MODERATE
            else:
                o = STRONG
            raw[s, o] += 1

    # Report raw counts
    obs_labels = ['Poor', 'Weak', 'Mod.', 'Strong']
    state_labels_short = ['Low', 'Stable', 'Growth']
    print(f"  Raw counts (before Laplace-1 smoothing):")
    print(f"  {'':8s}  " + "  ".join(f"{l:>6}" for l in obs_labels) + "  Total")
    for s in range(N_STATES):
        r = raw[s]
        print(f"  {state_labels_short[s]:8s}: " +
              "  ".join(f"{v:6d}" for v in r) + f"  {r.sum():5d}")

    counts = raw + 1  # Laplace-1
    Omega = counts / counts.sum(axis=1, keepdims=True)
    return Omega


def build_observation_model(ucas_data, state_seqs, pct_changes):
    counts = np.ones((N_STATES, N_OBS))  # Laplace smoothing
    for inst in INST_NAMES:
        app_hist = [ucas_data[inst][yr][0] for yr in CYCLES if yr in ucas_data[inst]]
        plc_hist = [ucas_data[inst][yr][1] for yr in CYCLES if yr in ucas_data[inst]]
        for i, yr in enumerate(CYCLES):
            if yr not in ucas_data[inst]:
                continue
            apps, placed = ucas_data[inst][yr]
            window = app_hist[max(0, i-5):i]
            avg5 = np.mean(window) if len(window) >= 2 else apps
            ratio = apps / avg5 if avg5 > 0 else 1.0
            if i >= 1 and plc_hist[i-1] > 0:
                yield_curr = placed / apps
                yield_prev = plc_hist[i-1] / app_hist[i-1]
                yield_improving = yield_curr >= yield_prev
            else:
                yield_improving = True
            if ratio >= 1.05 and yield_improving:
                o = STRONG
            elif 0.95 <= ratio < 1.05:
                o = MODERATE
            elif 0.85 <= ratio < 0.95:
                o = WEAK
            else:
                o = POOR
            # Map UCAS cycle year to % change year index
            # UCAS cycle yr → enrolment outcome in same academic year
            # % change series covers ACAD_YEARS[1:] = '2015/16' to '2024/25'
            # UCAS 2015 → index 0 in PCT_YEARS
            try:
                pct_idx = PCT_YEARS.index(f"{yr}/{str(yr+1)[2:]}")
            except ValueError:
                continue
            if 0 <= pct_idx < len(state_seqs[inst]):
                s = state_seqs[inst][pct_idx]
                counts[s, o] += 1
    Omega = counts / counts.sum(axis=1, keepdims=True)
    return Omega


# ============================================================
# STEP 3 — HESA FINANCE (data_2/table-1.csv) → R_NET
# ============================================================

def load_hesa_finance():
    """Load HESA Finance Statistics from data_2/table-1.csv (2015/16–2023/24).
    Returns dict {inst: {acad_year: (total_income_k, surplus_k)}}.
    Pre-2015/16 years use hardcoded clean-window values from c_sum_11/12.
    """
    path = DATA_FILES["hesa_finance_csv"]
    data = {inst: {} for inst in INST_NAMES}

    if os.path.exists(path):
        with open(path, encoding='utf-8-sig') as f:
            lines = f.readlines()
        hdr_idx = next(i for i, l in enumerate(lines) if l.startswith('UKPRN'))
        reader = csv.DictReader(lines[hdr_idx:])
        income_tmp  = {inst: {} for inst in INST_NAMES}
        surplus_tmp = {inst: {} for inst in INST_NAMES}
        for row in reader:
            ukprn = str(row.get('UKPRN', '')).strip()
            if ukprn not in UKPRN_MAP:
                continue
            inst     = UKPRN_MAP[ukprn]
            category = row.get('Category', '').strip()
            acad_yr  = row.get('Academic year', '').strip()
            raw_val  = str(row.get('Value(£000s)', '') or '').strip()
            if not raw_val:
                continue
            if raw_val.startswith('(') and raw_val.endswith(')'):
                try: val = -float(raw_val[1:-1].replace(',', ''))
                except: continue
            else:
                try: val = float(raw_val.replace(',', ''))
                except: continue
            if category == 'Total income':
                income_tmp[inst][acad_yr] = val
            elif category == 'Surplus/(deficit) for the year':
                surplus_tmp[inst][acad_yr] = val
        for inst in INST_NAMES:
            for yr in income_tmp[inst]:
                if yr in surplus_tmp[inst]:
                    data[inst][yr] = (income_tmp[inst][yr], surplus_tmp[inst][yr])
        n_loaded = sum(len(v) for v in data.values())
        print(f"  [REAL DATA] Loaded HESA Finance Statistics from {path} ({n_loaded} inst-year pairs)")
    else:
        print(f"  [MISSING] {path} not found — R_net will use clean-window constants only")

    return data


def build_reward_matrix(finance_data, state_seqs):
    """
    Build R_net using:
    1. Surplus observations from finance_data matched to state_seqs (overlapping years)
    2. Pre-computed clean-window surplus% from CLEAN_SURPLUS_PCT (c_sum_12, 14 clean years)
    Mean and std are pooled across both sources; source 1 provides state-aligned validation.
    Reward scale: 1% surplus ≈ 8 reward units (calibrated to observed variance).
    """
    scale = 8.0

    # --- Collect surplus% from finance_data matched to state_seqs ---
    # state_seqs covers PCT_YEARS = ['2015/16', …, '2024/25']
    # Clean window: use only 2015/16, 2016/17, 2017/18 (pre-USS distortion)
    CLEAN_CSV_YEARS = {'2015/16', '2016/17', '2017/18'}
    matched_surplus = {s: [] for s in range(N_STATES)}
    for inst in INST_NAMES:
        for i, yr in enumerate(PCT_YEARS):
            if yr not in CLEAN_CSV_YEARS:
                continue
            if yr not in finance_data.get(inst, {}):
                continue
            ti, su = finance_data[inst][yr]
            if ti <= 0:
                continue
            surplus_pct = su / ti * 100
            if i < len(state_seqs[inst]):
                s = state_seqs[inst][i]
                matched_surplus[s].append(surplus_pct)

    # --- Pool with clean-window constants ---
    combined = {}
    for s in range(N_STATES):
        combined[s] = CLEAN_SURPLUS_PCT[s] + matched_surplus[s]

    mean_surplus = {s: float(np.mean(combined[s])) for s in range(N_STATES)}
    std_surplus  = {s: float(np.std(combined[s]))  for s in range(N_STATES)}

    print(f"  Pooled surplus% by regime (clean window + matched CSV years):")
    for s in range(N_STATES):
        n_clean = len(CLEAN_SURPLUS_PCT[s])
        n_match = len(matched_surplus[s])
        print(f"    {STATE_NAMES[s]:7s}: mean={mean_surplus[s]:+.2f}%  "
              f"std={std_surplus[s]:.2f}%  n={len(combined[s])} ({n_clean} clean + {n_match} matched)")

    r_maintain = np.array([mean_surplus[s] * scale for s in range(N_STATES)])

    R_net = np.array([
        r_maintain + np.array([ 12.0, -4.0, -18.0]),   # CONSERVE
        r_maintain,                                      # MAINTAIN
        r_maintain + np.array([ -5.0, -2.0,  +8.0]),   # RECRUIT
        r_maintain + np.array([-35.0, -12.0, +45.0]),  # INVEST
        r_maintain + np.array([ +20.0, -6.0, -32.0]),  # RATIONALISE
    ])
    C_action = np.array([1.5, 0, 6.0, 12.0, 2.5])
    R_net -= C_action[:, np.newaxis]

    return R_net, mean_surplus, std_surplus


# ============================================================
# STEP 4 — DEMOGRAPHIC PRIOR → b0 BASE
# ONS 2024-based National Population Projections, England, Ages 15-29
# Published 28 April 2026. Replaces discontinued UCAS Futures (c_sum_17).
# ============================================================

# Ages 15-29 (thousands) from ONS 2024-based NPP, England principal projection.
# Source: enpppsummary.xlsx, sheet PERSONS, Table 2, row 'Ages 15 to 29'.
_ONS_AGES1529 = {
    2024: 10920.744, 2025: 10938.477, 2026: 10942.549, 2027: 10972.617,
    2028: 10996.514, 2029: 11017.974, 2030: 11069.019, 2031: 11143.603,
    2032: 11202.653, 2033: 11242.196, 2034: 11258.645, 2035: 11251.608,
}

def _ons_cohort_growth_rates(horizon_years=10):
    """Return list of YoY % growth rates for Ages 15-29, 2025 to 2024+horizon."""
    rates = []
    for yr in range(2025, 2025 + horizon_years):
        if yr in _ONS_AGES1529 and yr - 1 in _ONS_AGES1529:
            rates.append(_ONS_AGES1529[yr] / _ONS_AGES1529[yr - 1] - 1)
    return rates

def load_demographic_prior():
    """
    Build b0 demographic prior from ONS 2024-based NPP.

    Method (c_sum_17; thresholds updated c_sum_18):
    - Extract Ages 15-29 YoY growth rates for 5-year (near-term) and 10-year windows.
    - Map mean growth rates to prior weights using the POMDP state thresholds as anchors:
        Low_thresh  = -0.0071  (33rd percentile of pooled HESA Δ%, c_sum_10)
        High_thresh = +0.0376  (67th percentile)
    - Positive cohort growth within the Stable regime → Stable dominant, slight Growth lean.
    - Reads from data_3/ons_npp_2024based_england_summary.xlsx if available;
      falls back to hardcoded _ONS_AGES1529 constants (already extracted).
    """
    path = DATA_FILES["ons_npp"]

    # Attempt to refresh from file (validates file is present and readable)
    if os.path.exists(path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb['PERSONS']
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            # Row 26 (0-indexed) = year headers; row 28 = Ages 15-29 values
            year_hdr = [v for v in rows[26][1:] if v is not None]
            vals_row = rows[28]
            label = vals_row[0]
            for i, (yr, val) in enumerate(zip(year_hdr, vals_row[1:])):
                try:
                    _ONS_AGES1529[int(yr)] = float(val)
                except (TypeError, ValueError):
                    pass
            print(f"  [REAL DATA] ONS 2024-based NPP loaded: {label}, "
                  f"England Ages 15-29 2024={_ONS_AGES1529[2024]:.1f}k "
                  f"→ 2034={_ONS_AGES1529[2034]:.1f}k")
        except Exception as e:
            print(f"  [WARNING] ONS NPP read error: {e}; using embedded constants")
    else:
        print(f"  [REAL DATA] ONS 2024-based NPP: using embedded constants (file not found at {path})")

    # Compute growth rates over 5-year and 10-year horizons
    rates_5yr  = _ons_cohort_growth_rates(5)
    rates_10yr = _ons_cohort_growth_rates(10)
    mean_5yr   = float(np.mean(rates_5yr))   # +0.178%/yr (2025-2029)
    mean_10yr  = float(np.mean(rates_10yr))  # +0.305%/yr (2025-2034)

    print(f"  ONS Ages 15-29 mean YoY: 5yr={mean_5yr:+.3%}  10yr={mean_10yr:+.3%}")

    # POMDP state thresholds (empirical, c_sum_10)
    low_t  = -0.0071
    high_t = +0.0376
    stable_range = high_t - low_t  # 0.0447

    # Map mean 10-year growth to a Growth-lean weight
    # 0% growth → no Growth lean; high_t growth → maximum Growth lean
    growth_lean = np.clip(mean_10yr / stable_range, 0.0, 1.0)  # 0.305%/4.47% = 6.8%

    # Base prior: Stable-dominant; redistribute growth_lean × 0.08 from Low to Growth
    base_low    = 0.22
    base_stable = 0.55
    base_growth = 0.23
    shift = growth_lean * 0.08   # max shift = 0.08 at high threshold growth; here ≈ 0.005
    prior = np.array([base_low - shift, base_stable, base_growth + shift])
    prior = np.clip(prior, 0.01, 0.99)
    prior /= prior.sum()

    print(f"  b0_demo derived: Low={prior[0]:.3f}  Stable={prior[1]:.3f}  Growth={prior[2]:.3f}  "
          f"(growth_lean={growth_lean:.3f})")
    return prior


# ============================================================
# STEP 5 — HOME OFFICE VISA STATS → INTERNATIONAL ADJUSTMENT
# ============================================================

def load_visa_adjustment():
    """
    Load real Home Office visa statistics from data_3/home_office_edu_visas_sep2025.xlsx.
    Computes annual degree-level grant trends (Bachelors + Masters + Doctoral).
    Returns institution-specific b0 adjustments derived from:
      - National visa_trend_t (c_sum_14 §10.2)
      - Institution overseas enrolment share (HESA domicile proxy, c_sum_14 §4.5)
    """
    path = DATA_FILES["ho_visas"]
    if os.path.exists(path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb['Data_Edu_D02']
            rows = list(ws.iter_rows(values_only=True))
            # rows[2] = header, rows[3:] = data
            annual_deg = {}
            for row in rows[3:]:
                year, quarter, nat, region, vtype, level, grants = row
                if year is None or grants is None:
                    continue
                try:
                    year = int(year); grants = int(grants)
                except (TypeError, ValueError):
                    continue
                if level in ('Bachelors', 'Masters', 'Doctoral'):
                    annual_deg[year] = annual_deg.get(year, 0) + grants
            wb.close()

            # Compute visa_trend_t for recent years
            visa_trends = {}
            for yr in sorted(annual_deg):
                if yr - 1 in annual_deg and annual_deg[yr-1] > 0:
                    visa_trends[yr] = annual_deg[yr] / annual_deg[yr-1] - 1

            print(f"  [REAL DATA] Loaded Home Office visa statistics from {path}")
            print(f"  Degree-level annual grants: " +
                  "  ".join(f"{yr}:{g:,}" for yr, g in sorted(annual_deg.items())))
            print(f"  visa_trend_t: " +
                  "  ".join(f"{yr}:{t:+.1%}" for yr, t in sorted(visa_trends.items())
                             if yr >= 2019 and yr != 2020))  # excl. COVID anomaly

            # Most recent calibration-usable trend (2023→2024: -14.0%)
            trend_2324 = visa_trends.get(2024, -0.14)
            trend_2223 = visa_trends.get(2023, -0.063)

            # Institution-specific overseas share weights (from HESA domicile data, c_sum_14)
            # Warwick: high Indian PGT + Chinese → most exposed
            # Leeds:   high Indian PGT + Nigerian → heavily exposed
            # Exeter:  more European + Chinese, lower Indian PGT → less exposed
            # Bristol: diverse income base → least exposed
            overseas_weight = {"Warwick": 0.38, "Exeter": 0.22, "Leeds": 0.35, "Bristol": 0.24}

            # Map visa_trend to b0 shift: negative trend → shift mass toward Low
            # Scale: overseas_weight × |trend| determines magnitude of shift
            adjustments = {}
            for inst in INST_NAMES:
                w = overseas_weight[inst]
                # 2-year trend (2022→2024 composite): use mean of -6.3% and -14.0%
                composite_trend = (trend_2223 + trend_2324) / 2   # ≈ -10.2%
                shift = w * abs(composite_trend) if composite_trend < 0 else 0.0
                adjustments[inst] = np.array([+shift, -shift * 0.6, -shift * 0.4])

            for inst in INST_NAMES:
                print(f"    {inst:8s} overseas_wt={overseas_weight[inst]:.2f}  "
                      f"b0_shift={adjustments[inst].round(3)}")
            return adjustments

        except Exception as e:
            print(f"  [WARNING] Could not parse {path}: {e}")

    print("  [SYNTHETIC] Using documented Home Office visa statistics (2023-24 trend: -14%)")
    adjustments = {
        "Warwick": np.array([+0.15, -0.10, -0.05]),
        "Exeter":  np.array([+0.03, -0.02, -0.01]),
        "Leeds":   np.array([+0.12, -0.08, -0.04]),
        "Bristol": np.array([+0.05, -0.03, -0.02]),
    }
    return adjustments


# ============================================================
# INTEGRATION: CONSTRUCT b0 PER INSTITUTION
# ============================================================

def build_initial_priors(demographic_prior, visa_adjustments):
    jan_2026_signal = {
        "Warwick": WEAK, "Exeter": MODERATE, "Leeds": WEAK, "Bristol": MODERATE,
    }
    signal_likelihoods = {
        STRONG:   np.array([0.05, 0.20, 0.75]),
        MODERATE: np.array([0.10, 0.60, 0.30]),
        WEAK:     np.array([0.40, 0.50, 0.10]),
        POOR:     np.array([0.70, 0.28, 0.02]),
    }
    b0 = {}
    for inst in INST_NAMES:
        b = demographic_prior.copy()
        b += visa_adjustments[inst]
        b = np.clip(b, 0.01, 1.0); b /= b.sum()
        likelihood = signal_likelihoods[jan_2026_signal[inst]]
        b_upd = likelihood * b
        b_upd /= b_upd.sum()
        b0[inst] = b_upd
    return b0, jan_2026_signal


# ============================================================
# POMDP CORE (SR1–SR5 + strategies + Monte Carlo)
# ============================================================

def sr1_belief_update(b, o, T, O):
    b_pred = T.T @ b
    b_upd  = O[:, o] * b_pred
    norm   = b_upd.sum()
    return b_upd / norm if norm > 1e-12 else b_pred / (b_pred.sum() + 1e-12)


def sr3_satisficing_prob(b, T, O, R_net, U_min, gamma, horizon, n_rollouts, rng):
    successes = 0
    for _ in range(n_rollouts):
        b_c = b.copy(); total = 0.0
        for t in range(horizon):
            a = int(np.argmax(R_net @ b_c))
            s = rng.choice(N_STATES, p=b_c)
            total += (gamma ** t) * R_net[a, s]
            o = rng.choice(N_OBS, p=O[s])
            b_c = sr1_belief_update(b_c, o, T, O)
        if total >= U_min:
            successes += 1
    return successes / n_rollouts


def sr4_myopic_voi(b, T, O, R_net):
    v_no_info = float(np.max(R_net @ b))
    v_with_info = 0.0
    for o in range(N_OBS):
        p_o = float(O[:, o] @ b)
        if p_o < 1e-10: continue
        b_upd = O[:, o] * b; b_upd /= b_upd.sum()
        v_with_info += p_o * float(np.max(R_net @ b_upd))
    return max(0.0, v_with_info - v_no_info)


def sr5_option_value(b, action_irrev, T, O, R_net, gamma, depth=3):
    v_commit = float(R_net[action_irrev] @ b)
    b_after = T.T @ b
    for d in range(1, depth):
        best_a = int(np.argmax(R_net @ b_after))
        v_commit += (gamma ** d) * float(R_net[best_a] @ b_after)
        b_after = T.T @ b_after
    reversible = [a for a in range(N_ACTIONS) if a not in [INVEST, RATIONALISE]]
    best_rev = max(reversible, key=lambda a: float(R_net[a] @ b))
    v_delay = float(R_net[best_rev] @ b)
    b_next = np.zeros(N_STATES)
    for o in range(N_OBS):
        p_o = float(O[:, o] @ (T.T @ b))
        if p_o < 1e-10: continue
        b_upd = O[:, o] * (T.T @ b); b_upd /= b_upd.sum()
        b_next += p_o * b_upd
    v_delay += gamma * float(R_net[int(np.argmax(R_net @ b_next))] @ b_next)
    return v_delay - v_commit


def strat_bdt_baseline(b, obs_hist, t, rng, **kw): return MAINTAIN
def strat_myopic_pomdp(b, obs_hist, t, rng, **kw): return int(np.argmax(kw["R_net"] @ b))
def strat_dapp(b, obs_hist, t, rng, **kw):
    if len(obs_hist) >= 3 and all(o <= POOR for o in obs_hist[-3:]): return RATIONALISE
    if len(obs_hist) >= 2 and all(o <= WEAK for o in obs_hist[-2:]): return CONSERVE
    return MAINTAIN
def strat_rdm_conservative(b, obs_hist, t, rng, **kw):
    if b[LOW] > 0.55: return RATIONALISE
    if b[LOW] > 0.35: return CONSERVE
    if b[GROWTH] > 0.55: return RECRUIT
    return MAINTAIN
def strat_robust_satisficing(b, obs_hist, t, rng, horizon=T_HORIZON, **kw):
    remaining = max(1, horizon - t)
    sp = sr3_satisficing_prob(b, kw["T"], kw["O"], kw["R_net"], U_MIN, GAMMA,
                               min(remaining, 6), N_SAT, rng)
    return CONSERVE if sp < (1 - DELTA) else int(np.argmax(kw["R_net"] @ b))
def strat_unified(b, obs_hist, t, rng, horizon=T_HORIZON, **kw):
    remaining = max(1, horizon - t)
    sp = sr3_satisficing_prob(b, kw["T"], kw["O"], kw["R_net"], U_MIN, GAMMA,
                               min(remaining, 6), N_SAT, rng)
    if sp < (1 - DELTA): return CONSERVE
    Q = kw["R_net"] @ b
    voi = sr4_myopic_voi(b, kw["T"], kw["O"], kw["R_net"])
    Q[RECRUIT] += LAMBDA_V * voi
    for a_irrev in [INVEST, RATIONALISE]:
        ov = sr5_option_value(b, a_irrev, kw["T"], kw["O"], kw["R_net"], GAMMA)
        if ov > 0: Q[a_irrev] -= MU_OV * ov
    return int(np.argmax(Q))


def simulate_episode(strategy_fn, T_nominal, T_true, O_agent, O_true,
                     R_net, b0, rng):
    b = b0.copy()
    s = rng.choice(N_STATES, p=b0)
    obs_hist, total_r, distress, acts, sts = [], 0.0, 0, [], []
    kwargs = {"T": T_nominal, "O": O_agent, "R_net": R_net}
    for t in range(T_HORIZON):
        a = strategy_fn(b, obs_hist, t, rng=rng, **kwargs)
        acts.append(a); sts.append(s)
        r = R_net[a, s]
        total_r += (GAMMA ** t) * r
        if r < PER_STEP_CRISIS: distress += 1
        s  = rng.choice(N_STATES, p=T_true[s])
        o  = rng.choice(N_OBS, p=O_true[s])
        obs_hist.append(o)
        b  = sr1_belief_update(b, o, T_nominal, O_agent)
    return total_r, distress, acts, sts


def run_monte_carlo(T_nominal, T_true, O_agent, b0_pooled, R_net,
                    n_trials=N_TRIALS, seed=42):
    strategies = [
        ("BDT Baseline",       strat_bdt_baseline),
        ("Robust Satisficing", strat_robust_satisficing),
        ("Myopic POMDP",       strat_myopic_pomdp),
        ("DAPP",               strat_dapp),
        ("RDM-Conservative",   strat_rdm_conservative),
        ("Unified Framework",  strat_unified),
    ]
    master_rng = np.random.default_rng(seed)
    results = {}
    for name, fn in strategies:
        rewards, distress_rates, adists = [], [], []
        trial_rng = np.random.default_rng(master_rng.integers(10**9))
        for _ in range(n_trials):
            ep_rng = np.random.default_rng(trial_rng.integers(10**9))
            r, nd, acts, _ = simulate_episode(fn, T_nominal, T_true, O_agent,
                                               O_agent, R_net, b0_pooled, ep_rng)
            rewards.append(r); distress_rates.append(nd / T_HORIZON)
            ac = np.zeros(N_ACTIONS)
            for a in acts: ac[a] += 1
            adists.append(ac / len(acts))
        results[name] = {
            "failure_rate":     float(np.mean(distress_rates)),
            "mean_reward":      float(np.mean(rewards)),
            "std_reward":       float(np.std(rewards)),
            "ci_95_low":        float(np.percentile(rewards, 2.5)),
            "ci_95_high":       float(np.percentile(rewards, 97.5)),
            "mean_action_dist": np.mean(adists, axis=0),
        }
        r = results[name]
        print(f"  {name:<22}  distress={r['failure_rate']*100:5.1f}%  "
              f"reward={r['mean_reward']:7.1f}  "
              f"[{r['ci_95_low']:7.1f}, {r['ci_95_high']:7.1f}]")
    return results


# ============================================================
# FIGURES
# ============================================================

def plot_enrollment_regimes(pct_changes, state_seqs, low_t, high_t):
    fig, axes = plt.subplots(2, 2, figsize=(14, 8), sharex=True)
    colors_s = ['#d62728', '#1f77b4', '#2ca02c']
    labels_s = ['Low', 'Stable', 'Growth']
    for ax, inst in zip(axes.flatten(), INST_NAMES):
        delta  = pct_changes[inst]
        states = state_seqs[inst]
        ts     = np.array(YEARS[1:len(delta)+1])   # % change years: 2015…2024
        for s, c, lbl in zip([LOW, STABLE, GROWTH], colors_s, labels_s):
            mask = states == s
            ax.scatter(ts[mask], delta[mask], color=c, label=lbl, s=40, zorder=3)
        ax.bar(ts, delta, color='lightgrey', alpha=0.5, width=0.7, zorder=1)
        ax.axhline(low_t,  color='#d62728', linestyle='--', linewidth=0.8, alpha=0.7)
        ax.axhline(high_t, color='#2ca02c', linestyle='--', linewidth=0.8, alpha=0.7)
        ax.axhline(0, color='k', linewidth=0.5)
        ax.set_title(inst, fontsize=11, fontweight='bold')
        ax.set_ylabel("Annual % change", fontsize=9)
        if ax == axes[0, 0]: ax.legend(fontsize=8)
    fig.suptitle("Enrolment Regime Classification — Russell Group Panel (2015–2024)\n"
                 "REAL HESA data  |  Dashed lines: 33rd/67th percentile thresholds",
                 fontsize=11, fontweight='bold')
    plt.tight_layout(rect=[0, 0, 1, 0.93])
    plt.savefig("rg_enrollment_regimes.png", dpi=150, bbox_inches='tight')
    print("  Figure saved: rg_enrollment_regimes.png")


def plot_calibrated_params(T_pooled, Omega, R_net, b0):
    fig, axes = plt.subplots(1, 3, figsize=(15, 4))
    ax = axes[0]
    im = ax.imshow(T_pooled, vmin=0, vmax=1, cmap='Blues')
    ax.set_xticks(range(3)); ax.set_xticklabels(STATE_NAMES)
    ax.set_yticks(range(3)); ax.set_yticklabels(STATE_NAMES)
    ax.set_title("T (real HESA HMM)", fontweight='bold')
    ax.set_xlabel("To state"); ax.set_ylabel("From state")
    for i in range(3):
        for j in range(3):
            ax.text(j, i, f"{T_pooled[i,j]:.2f}", ha='center', va='center',
                    fontsize=10, color='white' if T_pooled[i,j] > 0.5 else 'black')
    plt.colorbar(im, ax=ax, fraction=0.046)
    ax = axes[1]
    im2 = ax.imshow(Omega, vmin=0, vmax=0.8, cmap='Oranges')
    ax.set_xticks(range(4)); ax.set_xticklabels(['Poor','Weak','Mod.','Strong'], fontsize=9)
    ax.set_yticks(range(3)); ax.set_yticklabels(STATE_NAMES)
    ax.set_title("Ω (synthetic UCAS proxy)", fontweight='bold')
    ax.set_xlabel("Signal"); ax.set_ylabel("True state")
    for i in range(3):
        for j in range(4):
            ax.text(j, i, f"{Omega[i,j]:.2f}", ha='center', va='center',
                    fontsize=9, color='white' if Omega[i,j] > 0.4 else 'black')
    plt.colorbar(im2, ax=ax, fraction=0.046)
    ax = axes[2]
    im3 = ax.imshow(R_net, cmap='RdYlGn', vmin=-50, vmax=80)
    ax.set_xticks(range(3)); ax.set_xticklabels(STATE_NAMES)
    ax.set_yticks(range(5)); ax.set_yticklabels(ACTION_NAMES, fontsize=8)
    ax.set_title("R_net (real finance clean window)", fontweight='bold')
    ax.set_xlabel("Demand state"); ax.set_ylabel("Action")
    for i in range(5):
        for j in range(3):
            ax.text(j, i, f"{R_net[i,j]:.0f}", ha='center', va='center',
                    fontsize=9, color='black')
    plt.colorbar(im3, ax=ax, fraction=0.046)
    fig.suptitle("Calibrated POMDP Parameters — Russell Group Panel  [c_sum_18]",
                 fontsize=12, fontweight='bold')
    plt.tight_layout(rect=[0, 0, 1, 0.93])
    plt.savefig("rg_calibrated_params.png", dpi=150, bbox_inches='tight')
    print("  Figure saved: rg_calibrated_params.png")


def plot_mc_comparison(results):
    names = list(results.keys())
    fail  = [results[n]["failure_rate"] * 100 for n in names]
    rew   = [results[n]["mean_reward"] for n in names]
    ci_lo = [rew[i] - results[names[i]]["ci_95_low"]  for i in range(len(names))]
    ci_hi = [results[names[i]]["ci_95_high"] - rew[i] for i in range(len(names))]
    colors = ['#d62728','#ff7f0e','#bcbd22','#2ca02c','#1f77b4','#9467bd']
    y = np.arange(len(names))
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
    bars = ax1.barh(y, fail, color=colors, alpha=0.85, edgecolor='k', linewidth=0.5)
    ax1.set_yticks(y); ax1.set_yticklabels(names, fontsize=10)
    ax1.set_xlabel("Distress-year rate (%)", fontsize=10)
    ax1.set_title("Financial Distress Rate  [real parameters]\n(% years reward < threshold)", fontsize=10)
    for bar, v in zip(bars, fail):
        ax1.text(bar.get_width()+0.2, bar.get_y()+bar.get_height()/2,
                 f"{v:.1f}%", va='center', fontsize=9)
    ax2.barh(y, rew, xerr=[ci_lo, ci_hi], color=colors, alpha=0.85,
             edgecolor='k', linewidth=0.5, capsize=4)
    ax2.set_yticks(y); ax2.set_yticklabels(names, fontsize=10)
    ax2.set_xlabel("Mean total discounted reward", fontsize=10)
    ax2.set_title("Mean Total Discounted Reward  [real parameters]\n(with 95% CI)", fontsize=10)
    ax2.axvline(0, color='k', linewidth=0.8)
    fig.suptitle(f"Russell Group POMDP — Six-Strategy Comparison  [c_sum_18]\n"
                 f"(n={N_TRIALS}, T={T_HORIZON} yrs, γ={GAMMA}, real calibrated parameters)",
                 fontsize=11, fontweight='bold')
    plt.tight_layout(rect=[0, 0, 1, 0.92])
    plt.savefig("rg_mc_comparison.png", dpi=150, bbox_inches='tight')
    print("  Figure saved: rg_mc_comparison.png")


def plot_b0_beliefs(b0):
    fig, ax = plt.subplots(figsize=(8, 4))
    x = np.arange(N_INST); w = 0.25
    colors_s = ['#d62728', '#1f77b4', '#2ca02c']
    for si, (state, color) in enumerate(zip(STATE_NAMES, colors_s)):
        vals = [b0[inst][si] for inst in INST_NAMES]
        ax.bar(x + si*w - w, vals, width=w, label=state, color=color, alpha=0.8)
    ax.set_xticks(x); ax.set_xticklabels(INST_NAMES, fontsize=11)
    ax.set_ylabel("Probability", fontsize=10)
    ax.set_title("Institution-Specific Initial Priors b₀  [c_sum_18]\n"
                 "(demographic prior + real HO visa adjustment + Jan 2026 signal)",
                 fontsize=11, fontweight='bold')
    ax.legend(fontsize=9); ax.set_ylim(0, 0.85)
    ax.axhline(1/3, color='grey', linestyle=':', linewidth=0.8)
    plt.tight_layout()
    plt.savefig("rg_initial_priors.png", dpi=150, bbox_inches='tight')
    print("  Figure saved: rg_initial_priors.png")


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    print("=" * 65)
    print("Recalibrated POMDP — Russell Group Panel  [c_sum_18]")
    print("Real data: HESA enrolments + Finance Stats + HO visas + ONS NPP + HESA Ω")
    print("=" * 65)

    # ─── Step 1: HESA Table 1 → Empirical thresholds + T_empirical
    print("\n── Step 1: HESA Table 1 (real) → Empirical threshold → T_pooled")
    print("  Method: 33rd/67th percentile thresholds + Laplace-smoothed count matrix")
    print("  (GaussianHMM found degenerate local minimum on 10-transition panel; empirical")
    print("   threshold approach is transparent, reproducible, and matches c_sum_10.)")
    enrolments  = load_hesa_enrolments()
    pct_changes = compute_pct_change(enrolments)
    low_t, high_t = compute_empirical_thresholds(pct_changes)
    print(f"  Pooled % change thresholds: Low < {low_t:.2f}%,  Growth > {high_t:.2f}%")

    # Empirical state assignments
    state_seqs = assign_states_threshold(pct_changes, low_t, high_t)

    # Empirical T (Laplace-smoothed counts)
    T_pooled = compute_T_empirical(state_seqs)

    # HMM for regime means/variances (documentation only; state assignments not used)
    _, regime_means_hmm, regime_vars_hmm, _, _, _ = fit_pooled_hmm(pct_changes)
    # Recompute regime means from threshold assignments
    all_delta = np.concatenate(list(pct_changes.values()))
    all_states = np.concatenate(list(state_seqs.values()))
    regime_means = np.array([float(np.mean(all_delta[all_states == s]))
                              if np.any(all_states == s) else 0.0 for s in range(N_STATES)])
    regime_stds  = np.array([float(np.std(all_delta[all_states == s]))
                              if np.any(all_states == s) else 1.0 for s in range(N_STATES)])

    print(f"  Regime means: " +
          "  ".join(f"{STATE_NAMES[s]}={regime_means[s]:.2f}%" for s in range(N_STATES)))
    print(f"  Regime std:   " +
          "  ".join(f"{STATE_NAMES[s]}={regime_stds[s]:.2f}%" for s in range(N_STATES)))
    print(f"  Pooled T_pooled (empirical + Laplace-1):")
    for i, row in enumerate(T_pooled):
        print(f"    {STATE_NAMES[i]:8s}: {row.round(3)}")
    from numpy.linalg import eig
    eigvals, eigvecs = eig(T_pooled.T)
    idx = np.argmin(np.abs(eigvals - 1.0))
    pi_stat = np.abs(np.real(eigvecs[:, idx])); pi_stat /= pi_stat.sum()
    print(f"  Stationary: Low={pi_stat[0]:.3f}  Stable={pi_stat[1]:.3f}  Growth={pi_stat[2]:.3f}")
    n_trans = sum(len(s) for s in state_seqs.values())
    n_low   = int(np.sum(all_states == LOW))
    n_stable= int(np.sum(all_states == STABLE))
    n_growth= int(np.sum(all_states == GROWTH))
    print(f"  State counts: Low={n_low}  Stable={n_stable}  Growth={n_growth}  "
          f"(total {n_trans} across {N_INST} institutions)")
    state_labels = {LOW: 'L', STABLE: 'S', GROWTH: 'G'}
    print(f"\n  Regime assignments ({PCT_YEARS[0]}–{PCT_YEARS[-1]}):")
    for inst in INST_NAMES:
        labels = "".join(state_labels[s] for s in state_seqs[inst])
        pct_str = "  ".join(f"{v:+5.1f}" for v in pct_changes[inst])
        print(f"    {inst:8s}: {labels}  ({pct_str})")

    # ─── Step 1b: Extended panel T_pooled (7 institutions — c_sum_21)
    print("\n── Step 1b: Extended T_pooled  [7 institutions: +Durham, Nottingham, Sheffield]")
    enrolments_ext  = load_hesa_enrolments_extended()
    pct_changes_ext = compute_pct_change(enrolments_ext)
    state_seqs_ext  = assign_states_threshold(pct_changes_ext, low_t, high_t,
                                              inst_names=INST_NAMES_EXT)
    T_ext = compute_T_empirical(state_seqs_ext, inst_names=INST_NAMES_EXT)
    all_states_ext  = np.concatenate([state_seqs_ext[i] for i in INST_NAMES_EXT])
    n_trans_ext = sum(len(state_seqs_ext[i]) for i in INST_NAMES_EXT)
    print(f"  Transitions: {n_trans_ext}  ({len(INST_NAMES_EXT)} institutions × 10 years)")
    eigvals_e, eigvecs_e = np.linalg.eig(T_ext.T)
    idx_e = np.argmin(np.abs(eigvals_e - 1.0))
    pi_ext = np.abs(np.real(eigvecs_e[:, idx_e])); pi_ext /= pi_ext.sum()
    print(f"  Stationary: Low={pi_ext[0]:.3f}  Stable={pi_ext[1]:.3f}  Growth={pi_ext[2]:.3f}  "
          f"(4-inst: Low={pi_stat[0]:.3f} Stable={pi_stat[1]:.3f} Growth={pi_stat[2]:.3f})")
    print(f"  T_ext vs T_pooled max|Δ| = {np.max(np.abs(T_ext - T_pooled)):.3f}")
    print(f"  Extended T_pooled:")
    for i, row in enumerate(T_ext):
        print(f"    {STATE_NAMES[i]:8s}: {row.round(3)}")
    state_counts_ext = {s: int(np.sum(all_states_ext == s)) for s in range(N_STATES)}
    print(f"  State counts: Low={state_counts_ext[LOW]}  "
          f"Stable={state_counts_ext[STABLE]}  Growth={state_counts_ext[GROWTH]}")
    for inst in [i for i in INST_NAMES_EXT if i not in INST_NAMES]:
        labels = "".join({0:'L',1:'S',2:'G'}[s] for s in state_seqs_ext[inst])
        pct_str = "  ".join(f"{v:+5.1f}" for v in pct_changes_ext[inst])
        print(f"    {inst:10s}: {labels}  ({pct_str})")

    # ─── Step 2: HESA-based Ω (real data; UCAS path activated when account obtained)
    ucas_path = DATA_FILES["ucas_eoc"]
    if os.path.exists(ucas_path):
        print("\n── Step 2: UCAS EOC → Omega  [REAL — UCAS provider data]")
        ucas_data = load_ucas_eoc()
        Omega = build_observation_model(ucas_data, state_seqs, pct_changes)
    else:
        print("\n── Step 2: HESA-based Ω  [REAL — enrolment % change thresholds (c_sum_18)]")
        print(f"  Thresholds: Poor<0%  |  Weak∈[0%,{low_t:.2f}%)  |  "
              f"Moderate∈[{low_t:.2f}%,{high_t:.2f}%)  |  Strong≥{high_t:.2f}%")
        Omega = build_omega_hesa(pct_changes, state_seqs, low_t, high_t)
    print(f"  Omega[state, signal]:")
    print(f"  {'':8s}  {'Poor':>6}  {'Weak':>6}  {'Mod.':>6}  {'Strong':>6}")
    for s in range(N_STATES):
        print(f"  {STATE_NAMES[s]:8s}: " + "  ".join(f"{Omega[s,o]:6.3f}" for o in range(N_OBS)))

    # ─── Step 3: HESA Finance → R_net (real clean window)
    print("\n── Step 3: HESA Finance Statistics (real) → R_net")
    print("  Clean calibration window: 2004/05–2017/18 (14 years, pre-USS distortion)")
    finance_data = load_hesa_finance()
    R_net, mean_surp, std_surp = build_reward_matrix(finance_data, state_seqs)
    print(f"\n  R_net (actions × states):")
    print(f"  {'':13s}  {'Low':>7}  {'Stable':>7}  {'Growth':>7}")
    for a in range(N_ACTIONS):
        print(f"  {ACTION_NAMES[a]:13s}: " + "  ".join(f"{R_net[a,s]:7.1f}" for s in range(N_STATES)))

    # ─── Step 4: Demographic prior (synthetic)
    print("\n── Step 4: Demographic Prior  [REAL — ONS 2024-based NPP, Ages 15-29]")
    demographic_prior = load_demographic_prior()
    print(f"  b0_demo (Low, Stable, Growth): {demographic_prior.round(3)}")

    # ─── Step 5: Visa adjustment (real)
    print("\n── Step 5: Home Office Visa Stats (real) → International Adjustment")
    visa_adj = load_visa_adjustment()

    # ─── Integration: b0 per institution
    print("\n── Integration: Institution-Specific b₀")
    b0_inst, jan_signals = build_initial_priors(demographic_prior, visa_adj)
    for inst in INST_NAMES:
        sig_name = ["Poor","Weak","Moderate","Strong"][jan_signals[inst]]
        print(f"  {inst:8s}  Jan2026={sig_name:8s}  b0={b0_inst[inst].round(3)}")
    b0_pooled = np.mean(list(b0_inst.values()), axis=0)
    b0_pooled /= b0_pooled.sum()
    print(f"\n  Pooled b0: {b0_pooled.round(3)}")

    # ─── T_true (slight perturbation for MC robustness check)
    T_true = T_pooled.copy()
    T_true += np.random.default_rng(99).normal(0, 0.015, T_true.shape)
    T_true = np.clip(T_true, 0.01, 1); T_true /= T_true.sum(axis=1, keepdims=True)

    # ─── Monte Carlo (Step 6)
    print(f"\n── Step 6: Monte Carlo (n={N_TRIALS}, T={T_HORIZON}, 6 strategies)")
    print(f"  U_min={U_MIN}, distress_threshold={PER_STEP_CRISIS}\n")
    mc_results = run_monte_carlo(T_pooled, T_true, Omega, b0_pooled,
                                  R_net, n_trials=N_TRIALS, seed=42)

    # ─── Results table
    print("\n" + "=" * 90)
    print(f"{'Strategy':<22} {'Distress%':>10} {'Mean Reward':>12} {'95% CI':>22}")
    print("=" * 90)
    for name, res in mc_results.items():
        ci = f"[{res['ci_95_low']:7.1f}, {res['ci_95_high']:7.1f}]"
        print(f"{name:<22} {res['failure_rate']*100:>9.1f}% {res['mean_reward']:>12.1f}  {ci}")
    print("=" * 90)

    print("\n  Action usage (% of decision steps):")
    print(f"  {'Strategy':<22} " + " ".join(f"{n:>11}" for n in ACTION_NAMES))
    for name, res in mc_results.items():
        row = f"  {name:<22} " + " ".join(f"{v*100:>10.1f}%" for v in res["mean_action_dist"])
        print(row)

    # ─── Figures
    print("\n── Generating figures")
    plot_enrollment_regimes(pct_changes, state_seqs, low_t, high_t)
    plot_calibrated_params(T_pooled, Omega, R_net, b0_pooled)
    plot_mc_comparison(mc_results)
    plot_b0_beliefs(b0_inst)

    # ─── Comparison: synthetic vs real calibration
    # Synthetic reference from previous rg_pomdp_calibration.py runs (c_sum_8)
    print("\n── Comparison: Synthetic (c_sum_8) vs Real calibration (c_sum_18)")
    synthetic_ref = {
        "BDT Baseline":       (36.2,  40.0),
        "Robust Satisficing": ( 9.1,  91.6),
        "Myopic POMDP":       (10.1,  87.0),
        "DAPP":               (23.3,  60.5),
        "RDM-Conservative":   ( 7.8,  76.1),
        "Unified Framework":  (11.9,  91.2),
    }
    print(f"\n  {'Strategy':<22} {'Synth.Fail%':>12} {'Real Fail%':>11} "
          f"{'Synth.Rew':>11} {'Real Rew':>10} {'Δ Reward':>10}")
    print("  " + "-" * 80)
    for name in mc_results:
        sf, sr = synthetic_ref.get(name, (0, 0))
        rf  = mc_results[name]["failure_rate"] * 100
        rr  = mc_results[name]["mean_reward"]
        dr  = rr - sr
        flag = " ← REVERSAL" if (sf < 15 and rf > sf + 10) or (sf > 20 and rf < sf - 10) else ""
        print(f"  {name:<22} {sf:>11.1f}% {rf:>10.1f}%  {sr:>10.1f} {rr:>10.1f} {dr:>+10.1f}{flag}")

    # ─── Key parameter comparison table
    print("\n── Key parameter changes: Synthetic → Real")
    print(f"  {'Parameter':<30} {'Synthetic':>16} {'Real':>16} {'Change'}")
    print(f"  {'-'*70}")
    print(f"  {'T(Low→Low)':<30} {'~0.000 (degenerate)':>16} {T_pooled[LOW,LOW]:>16.3f}  CRITICAL")
    print(f"  {'T(Growth→Growth)':<30} {'~0.900':>16} {T_pooled[GROWTH,GROWTH]:>16.3f}")
    print(f"  {'π(Low) stationary':<30} {'~7%':>16} {pi_stat[0]*100:>15.1f}%")
    print(f"  {'π(Growth) stationary':<30} {'~50%':>16} {pi_stat[2]*100:>15.1f}%")
    print(f"  {'R_net MAINTAIN Low':<30} {'~29 units':>16} {R_net[MAINTAIN,LOW]:>15.1f}  units")
    print(f"  {'R_net MAINTAIN Stable':<30} {'~33 units':>16} {R_net[MAINTAIN,STABLE]:>15.1f}  units")
    print(f"  {'R_net MAINTAIN Growth':<30} {'~44 units':>16} {R_net[MAINTAIN,GROWTH]:>15.1f}  units")
    print(f"  {'Surplus Low (clean)':<30} {'N/A':>16} {mean_surp[LOW]:>15.2f}%")
    print(f"  {'Surplus Stable (clean)':<30} {'N/A':>16} {mean_surp[STABLE]:>15.2f}%")
    print(f"  {'Surplus Growth (clean)':<30} {'N/A':>16} {mean_surp[GROWTH]:>15.2f}%")

    print("\nDone.  Output: c_sum_18")
